import datetime
from io import BytesIO
from typing import List
from pydantic import BaseModel # Use pydantic.BaseModel directly
from fastapi import APIRouter, HTTPException, Depends, status, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
import json # for sheet_data loading
import re # for file_name_user_part sanitization
import base64 # For Base64 encoding
import zipfile # For creating ZIP files
import os # For os.path.splitext
import io # For BytesIO and MediaIoBaseDownload

from app.auth import AuthorizedUser # For user authentication

# Google API Client libraries
from google.oauth2.credentials import Credentials
import os
from app.env import Mode, mode # For Databutton fallback logic
from google.auth.transport.requests import Request as GoogleAuthRequest # Renamed to avoid conflict
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.errors import HttpError


import databutton as db
import json # for sheet_data loading
import os # For os.path.splitext and environment variables
import re # for file_name_user_part sanitization
import base64 # For Base64 encoding
import zipfile # For creating ZIP files
import io # For BytesIO and MediaIoBaseDownload

from app.auth import AuthorizedUser # For user authentication
from app.env import Mode, mode # For Databutton fallback logic

# Google API Client libraries
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleAuthRequest # Renamed to avoid conflict
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.errors import HttpError

# Firebase Admin SDK
import firebase_admin
from firebase_admin import credentials as firebase_credentials, firestore

# --- Firebase Initialization (local to this module) ---
firebase_admin_initialized_local = False
db_firestore_admin_client_local = None

def initialize_firebase_admin_local():
    global firebase_admin_initialized_local, db_firestore_admin_client_local
    if firebase_admin_initialized_local:
        print("INFO (export_service): Firebase Admin SDK already initialized locally.")
        return

    try:
        # Try to get service account key from environment variable first
        cred_json_str = os.environ.get("FIREBASE_SERVICE_ACCOUNT_KEY")
        source = "environment variable"

        if not cred_json_str:
            # Fallback to Databutton secrets if running in DEV mode or env var not set
            print("INFO (export_service): FIREBASE_SERVICE_ACCOUNT_KEY not in env, trying db.secrets...")
            if mode == Mode.DEV or os.environ.get("DATABUTTON_PROJECT_ID"): # Check if in Databutton context
                try:
                    cred_json_str = db.secrets.get("FIREBASE_SERVICE_ACCOUNT_KEY")
                    source = "db.secrets"
                    if cred_json_str:
                        print("INFO (export_service): FIREBASE_SERVICE_ACCOUNT_KEY loaded from db.secrets.")
                except Exception as e:
                    print(f"ERROR (export_service): Could not load FIREBASE_SERVICE_ACCOUNT_KEY from db.secrets: {e}")
            else:
                print("INFO (export_service): Not in Databutton DEV mode or context, db.secrets not checked for FIREBASE_SERVICE_ACCOUNT_KEY.")

        if cred_json_str:
            cred_json = json.loads(cred_json_str)
            cred = firebase_credentials.Certificate(cred_json)
            # Check if the default Firebase app is already initialized
            app_already_exists = False
            try:
                firebase_admin.get_app('[DEFAULT]')
                app_already_exists = True
                print(f"INFO (export_service): Firebase Admin SDK already initialized globally, re-using for export_service.")
            except firebase_admin.AppError: # Or a more specific exception if known for "not found"
                print(f"INFO (export_service): Default Firebase app not found, export_service will initialize.")
                pass # App doesn't exist, proceed to initialize
            
            if not app_already_exists:
                 firebase_admin.initialize_app(cred, name='[DEFAULT]') # Explicitly name it if creating
                 print(f"INFO (export_service): Firebase Admin SDK initialized by export_service from {source} as [DEFAULT].")
            
            db_firestore_admin_client_local = firestore.client()
            firebase_admin_initialized_local = True
            print("INFO (export_service): Firestore client obtained locally by export_service and Firebase Admin is initialized.")
        else:
            print("ERROR (export_service): FIREBASE_SERVICE_ACCOUNT_KEY not found in env or db.secrets. Firebase Admin SDK for export_service NOT initialized.")

    except Exception as e:
        print(f"ERROR (export_service): Failed to initialize Firebase Admin SDK or Firestore client for export_service: {e}")
        firebase_admin_initialized_local = False
        db_firestore_admin_client_local = None

initialize_firebase_admin_local() # Initialize on module load


# Define the router at the top level of the module
router = APIRouter(
    prefix="/export",
    tags=["Export Service"]
)

# --- Helper function for storage keys ---
def sanitize_storage_key(key: str) -> str:
    """Sanitize storage key to only allow alphanumeric and ._- symbols"""
    return re.sub(r'[^a-zA-Z0-9._-]', '', key)

# --- Endpoint to Upload Excel Template ---
class TemplateUploadResponse(BaseModel): # Changed from db.Pydantic
    message: str
    file_name: str
    size: int

@router.post("/upload-excel-template", response_model=TemplateUploadResponse)
async def upload_excel_template(file: UploadFile):
    """Receives an Excel file and stores it as the user's template."""
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No file name provided.")

    if not file.filename.endswith(('.xlsx')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid file type. Please upload an .xlsx file.")

    template_storage_key = "user_expense_template.xlsx"
    print(f"[TEMPLATE_UPLOAD] Attempting to upload template: {file.filename} to {template_storage_key}")
    
    try:
        contents = await file.read()
        db.storage.binary.put(template_storage_key, contents)
        print(f"[TEMPLATE_UPLOAD] Successfully uploaded {file.filename} ({len(contents)} bytes) as {template_storage_key}")
        return TemplateUploadResponse(
            message="Plantilla guardada correctamente.",
            file_name=file.filename,
            size=len(contents)
        )
    except Exception as e:
        print(f"[TEMPLATE_UPLOAD] Error uploading template {file.filename}: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Could not save template: {e}") from e

# Attempt to import models from the expense_api. 
try:
    from app.apis.expense_api import ExpenseSheet as ActualExpenseSheet, ExpenseEntry as ActualExpenseEntry, get_expense_sheet_storage_key
    print("[INFO_EXPORT_SERVICE] Successfully imported ActualExpenseSheet, ActualExpenseEntry, and get_expense_sheet_storage_key from app.apis.expense_api")
except ImportError:
    print("[WARNING_EXPORT_SERVICE] Failed to import from app.apis.expense_api. Using placeholder models/functions. This may lead to inconsistencies.")
    class ActualExpenseEntry(BaseModel): # Changed from db.Pydantic
        id: str
        entry_date: datetime.date | None = None
        merchant_name: str | None = None
        payment_method: str | None = None
        project: str | None = None
        company: str | None = None
        location: str | None = None
        receipt_google_drive_id: str | None = None
        receipt_google_drive_web_view_link: str | None = None
        receipt_google_drive_web_content_link: str | None = None
        receipt_google_drive_file_name: str | None = None
        parking_amount: float | None = None
        taxi_amount: float | None = None
        transport_amount: float | None = None
        hotel_amount: float | None = None
        lunch_amount: float | None = None
        dinner_amount: float | None = None
        miscellaneous_amount: float | None = None
        kilometers: float | None = None
        km_rate: float | None = None
        km_amount: float | None = None
        daily_total: float | None = None
        created_at: datetime.datetime
        updated_at: datetime.datetime

    class ActualExpenseSheet(BaseModel): # Changed from db.Pydantic
        id: str
        name: str
        month: int
        year: int
        currency: str
        payment_method_filter: str | None = None
        status: str
        comments: str | None = None
        user_name: str | None = None
        anticipo: float = 0.0
        total_amount: float = 0.0
        entries: List[ActualExpenseEntry] = []
        created_at: datetime.datetime
        updated_at: datetime.datetime

    def get_expense_sheet_storage_key(sheet_id: str) -> str:
        s_key = sanitize_storage_key(f"expense_sheet_{sheet_id}.json")
        print(f"[DEBUG_EXPORT_PLACEHOLDER] Generated storage key with placeholder: {s_key} for sheet ID: {sheet_id}")
        return s_key

def get_month_name(month_number: int) -> str:
    """Returns the Spanish name for a given month number."""
    months_es = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    try:
        return months_es[month_number - 1]
    except IndexError:
        return "Mes Desconocido"

def format_currency_cell(cell: Cell, currency_code="EUR"):
    if currency_code == "EUR":
        cell.number_format = '#,##0.00 €'
    elif currency_code == "USD":
        cell.number_format = '$#,##0.00'
    elif currency_code == "GBP":
        cell.number_format = '£#,##0.00'
    else:
        cell.number_format = f'#,##0.00 "{currency_code}"'

def set_cell_style(cell: Cell, bold=False, italic=False, alignment=None, fill=None, font_color=None, border=None, font_name='Calibri', font_size=11):
    cell.font = Font(name=font_name, size=font_size, bold=bold, italic=italic, color=font_color)
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border

class ExportSheetRequest(BaseModel): # Changed from db.Pydantic
    sheet_id: str

class ExportResponse(BaseModel): # Changed from db.Pydantic
    file_name: str
    file_content_base64: str

@router.post("/expense-sheet/export-excel", response_model=ExportResponse)
async def export_expense_sheet_to_excel(request_body: ExportSheetRequest) -> ExportResponse:
    sheet_id = request_body.sheet_id
    print(f"[BASE64_EXPORT_DEBUG] Entered export_expense_sheet_to_excel for sheet_id: {sheet_id}")
    try:
        storage_key = get_expense_sheet_storage_key(sheet_id)
        print(f"[BASE64_EXPORT_DEBUG] Attempting to load sheet from storage key: {storage_key}")
        sheet_data_from_storage = db.storage.json.get(storage_key, default=None)
        
        if sheet_data_from_storage is None:
            print(f"[BASE64_EXPORT_DEBUG] Expense sheet {sheet_id} not found at {storage_key}")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Expense sheet {sheet_id} not found")

        print(f"[BASE64_EXPORT_DEBUG] Data type from storage for {sheet_id}: {type(sheet_data_from_storage)}")

        loaded_dict = None
        if isinstance(sheet_data_from_storage, str):
            try:
                loaded_dict = json.loads(sheet_data_from_storage)
                print(f"[BASE64_EXPORT_DEBUG] Successfully json.loads'd string data for {sheet_id}")
            except json.JSONDecodeError as json_err:
                print(f"[BASE64_EXPORT_DEBUG] JSONDecodeError for sheet {sheet_id} when parsing string data: {json_err}. Data was: {sheet_data_from_storage}")
                raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error decoding JSON string data: {json_err}") from json_err
        elif isinstance(sheet_data_from_storage, dict):
            loaded_dict = sheet_data_from_storage
            print(f"[BASE64_EXPORT_DEBUG] Used dict data directly for {sheet_id}")
        else:
            print(f"[BASE64_EXPORT_DEBUG] Unexpected data type from storage for {sheet_id}: {type(sheet_data_from_storage)}. Data: {sheet_data_from_storage}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Unexpected data type from storage: {type(sheet_data_from_storage)}")

        if loaded_dict is None: # Should not happen if logic above is correct, but as a safeguard
             print(f"[BASE64_EXPORT_DEBUG] loaded_dict is None for sheet {sheet_id} after attempting to process from storage. This is unexpected.")
             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to load sheet data into a dictionary.")

        try:
            sheet = ActualExpenseSheet(**loaded_dict)
            print(f"[BASE64_EXPORT_DEBUG] Successfully parsed/validated sheet {sheet_id} with {len(sheet.entries)} entries using ActualExpenseSheet.")
        except Exception as parse_exc: 
            print(f"[BASE64_EXPORT_DEBUG] Error parsing/validating sheet {sheet_id} with ActualExpenseSheet: {parse_exc}. Data dictionary was: {loaded_dict}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error parsing/validating sheet data with Pydantic model: {parse_exc}") from parse_exc

        creator_first_name_to_write = "N/A"
        creator_last_name_to_write = "N/A"
        if sheet.user_id and firebase_admin_initialized_local and db_firestore_admin_client_local:
            try:
                profile_ref = db_firestore_admin_client_local.collection("user_profiles").document(sheet.user_id)
                profile_doc = profile_ref.get()
                if profile_doc.exists:
                    profile_data = profile_doc.to_dict()
                    creator_first_name_to_write = profile_data.get("firstName", "N/A")
                    creator_last_name_to_write = profile_data.get("lastName", "N/A")
                    print(f"[EXPORT_SERVICE_FIRESTORE] Fetched profile for {sheet.user_id}: First: {creator_first_name_to_write}, Last: {creator_last_name_to_write}")
                else:
                    print(f"[EXPORT_SERVICE_FIRESTORE] Profile document not found for user_id: {sheet.user_id} in 'user_profiles' collection.")
            except Exception as e:
                print(f"[EXPORT_SERVICE_FIRESTORE] Error fetching user profile from Firestore for user_id {sheet.user_id}: {e}")
        elif not sheet.user_id:
            print(f"[EXPORT_SERVICE_FIRESTORE] No user_id found in sheet {sheet_id}, cannot fetch creator name.")
        elif not firebase_admin_initialized_local or not db_firestore_admin_client_local:
            print(f"[EXPORT_SERVICE_FIRESTORE] Firestore client not initialized (local check), cannot fetch creator name for sheet {sheet_id}.")

        # Create a new workbook for the export, no template loading
        wb = Workbook()
        ws = wb.active
        # Sanitize sheet name for Excel (max 31 chars, no invalid chars)
        invalid_excel_sheet_chars = r'[\\/*?:\[\]]'
        sanitized_sheet_name = re.sub(invalid_excel_sheet_chars, '_', sheet.name)
        ws.title = sanitized_sheet_name[:31]
        print(f"[EXPORT_FIXED_V2] Creating new workbook for sheet: {sheet_id}, sheet title: '{ws.title}'")

        # -- Start of Custom Multi-Row Header --
        # Fila 1: Título Principal
        title_cell = ws.cell(row=1, column=1, value="INFORME DE GASTOS MENSUALES DEL PERSONAL")
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Ajustar al nuevo número de columnas de la tabla de datos (15 columnas: A-O)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15) 

        # Fila 2: Nombre y Apellidos
        ws.cell(row=2, column=1, value="NOMBRE Y APELLIDOS:").font = Font(bold=True)
        # Escribir firstName en C2 y lastName en D2 (valores obtenidos de Firestore)
        ws.cell(row=2, column=3, value=creator_first_name_to_write) # C2
        ws.cell(row=2, column=4, value=creator_last_name_to_write) # D2

        # Fila 3: Pago y Moneda
        ws.cell(row=3, column=1, value="PAGO:").font = Font(bold=True)
        ws.cell(row=3, column=3, value=sheet.payment_method_filter or "N/A")
        ws.cell(row=3, column=5, value="MONEDA:").font = Font(bold=True)
        ws.cell(row=3, column=7, value=sheet.currency or "N/A")

        # Fila 4: Mes y Año
        ws.cell(row=4, column=1, value="MES:").font = Font(bold=True)
        ws.cell(row=4, column=3, value=get_month_name(sheet.month) if sheet.month else "N/A")
        ws.cell(row=4, column=5, value="AÑO:").font = Font(bold=True)
        ws.cell(row=4, column=7, value=str(sheet.year) if sheet.year else "N/A")

        # Fila 5: Fila vacía para separación (opcional, pero mejora legibilidad)
        # La tabla de datos comenzará en la fila 6
        current_data_row_start = 6
        # -- End of Custom Multi-Row Header --

        # Define headers for the data table
        data_table_headers = [
            "FECHA", "PROYECTO", "EMPRESA", "LOCALIDAD", 
            "PARKING", "TAXI", "KM", "IMPORTE KMs", 
            "AVION/HOTEL/COCHE", "HOTEL", "ALMUERZO", "CENA", "VARIOS", 
            "TOTAL DIARIO", "Ticket Adjunto"
        ]
        # Append data table headers at current_data_row_start
        ws.append_row = current_data_row_start # type: ignore # openpyxl allows this assignment
        header_row_num = current_data_row_start
        for col_idx, header_title in enumerate(data_table_headers, 1):
            cell = ws.cell(row=header_row_num, column=col_idx, value=header_title)
            cell.font = Font(bold=True, color="FFFFFF") # White text
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue background
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin_border_side = Side(style='thin', color="000000")
            cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = len(header_title) + 5 if len(header_title) > 10 else 15

        # Data entries start from the row after data table headers
        data_entry_start_row = header_row_num + 1

        if not sheet.entries:
            print(f"[EXPORT_FIXED_V2] Sheet {sheet_id} has no entries to export.")
            empty_message_cell = ws.cell(row=data_entry_start_row, column=1, value="No hay gastos en esta hoja.")
            empty_message_cell.font = Font(italic=True)
            empty_message_cell.alignment = Alignment(horizontal="center")
            if len(data_table_headers) > 1:
                ws.merge_cells(start_row=data_entry_start_row, start_column=1, end_row=data_entry_start_row, end_column=len(data_table_headers))
        else:
            print(f"[EXPORT_FIXED_V2] Processing {len(sheet.entries)} entries for sheet {sheet_id}.")
            # Initialize accumulators for partial totals
            totals_accumulators = {
                "PARKING": 0.0,
                "TAXI": 0.0,
                "KM": 0.0, # Sum of kilometers themselves
                "IMPORTE KMs": 0.0,
                "AVION/HOTEL/COCHE": 0.0,
                "HOTEL": 0.0,
                "ALMUERZO": 0.0,
                "CENA": 0.0,
                "VARIOS": 0.0,
                "TOTAL DIARIO": 0.0 # Sum of all daily totals
            }

            for entry_idx, entry in enumerate(sheet.entries, start=data_entry_start_row):
                # Mapeo a las nuevas columnas:
                fecha_str = entry.entry_date.strftime("%d/%m/%Y") if entry.entry_date else ""
                proyecto_str = entry.project or ""
                empresa_str = entry.merchant_name or ""
                localidad_str = entry.location or ""
                
                parking_val = entry.parking_amount if entry.parking_amount is not None else 0.0
                totals_accumulators["PARKING"] += parking_val
                
                taxi_val = entry.taxi_amount if entry.taxi_amount is not None else 0.0
                totals_accumulators["TAXI"] += taxi_val
                
                km_val = entry.kilometers if entry.kilometers is not None else 0.0
                totals_accumulators["KM"] += km_val # Accumulating actual KMs
                
                importe_kms_val = entry.km_amount if entry.km_amount is not None else 0.0
                totals_accumulators["IMPORTE KMs"] += importe_kms_val
                
                avion_hotel_coche_val = entry.transport_amount if entry.transport_amount is not None else 0.0
                totals_accumulators["AVION/HOTEL/COCHE"] += avion_hotel_coche_val
                
                hotel_val = entry.hotel_amount if entry.hotel_amount is not None else 0.0
                totals_accumulators["HOTEL"] += hotel_val
                
                almuerzo_val = entry.lunch_amount if entry.lunch_amount is not None else 0.0
                totals_accumulators["ALMUERZO"] += almuerzo_val
                
                cena_val = entry.dinner_amount if entry.dinner_amount is not None else 0.0
                totals_accumulators["CENA"] += cena_val
                
                varios_val = entry.miscellaneous_amount if entry.miscellaneous_amount is not None else 0.0
                totals_accumulators["VARIOS"] += varios_val
                
                total_diario_val = entry.daily_total if entry.daily_total is not None else 0.0
                totals_accumulators["TOTAL DIARIO"] += total_diario_val

                # Check for associated ticket
                ticket_adjunto_val = "Sí" if entry.receipt_google_drive_id and entry.receipt_google_drive_file_name else "No"

                row_data_values = [
                    fecha_str, proyecto_str, empresa_str, localidad_str,
                    parking_val, taxi_val, km_val, importe_kms_val,
                    avion_hotel_coche_val, hotel_val, almuerzo_val, cena_val, varios_val,
                    total_diario_val, ticket_adjunto_val
                ]
                
                for col_idx_val, value in enumerate(row_data_values, 1):
                    cell = ws.cell(row=entry_idx, column=col_idx_val, value=value)
                    # Make TOTAL DIARIO column bold for data rows
                    if data_table_headers[col_idx_val -1] == "TOTAL DIARIO":
                        cell.font = Font(bold=True)
                    # Center align 'Ticket Adjunto' column
                    if data_table_headers[col_idx_val -1] == "Ticket Adjunto":
                        cell.alignment = Alignment(horizontal="center")
                
                # Apply thin border to all data cells
                data_table_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for col_idx_loop_data in range(1, len(data_table_headers) + 1):
                    data_cell = ws.cell(row=entry_idx, column=col_idx_loop_data)
                    data_cell.border = data_table_border

                ws.cell(row=entry_idx, column=data_table_headers.index("FECHA") + 1).alignment = Alignment(horizontal="left")
                km_cell_format = ws.cell(row=entry_idx, column=data_table_headers.index("KM") + 1)
                km_cell_format.number_format = '0.00'
                km_cell_format.alignment = Alignment(horizontal="right")

                currency_columns_names = ["PARKING", "TAXI", "IMPORTE KMs", "AVION/HOTEL/COCHE", "HOTEL", "ALMUERZO", "CENA", "VARIOS", "TOTAL DIARIO"]
                for col_name_currency in currency_columns_names:
                    currency_cell_format = ws.cell(row=entry_idx, column=data_table_headers.index(col_name_currency) + 1)
                    format_currency_cell(currency_cell_format, sheet.currency)
                    currency_cell_format.alignment = Alignment(horizontal="right")
                
                text_columns_to_align_left = ["PROYECTO", "EMPRESA", "LOCALIDAD"]
                for col_name_text_align in text_columns_to_align_left:
                    ws.cell(row=entry_idx, column=data_table_headers.index(col_name_text_align) + 1).alignment = Alignment(horizontal="left")
            
            # --- Add Summary Rows --- 
            current_row = ws.max_row + 1

            # Row: TOTALES PARCIALES
            ws.cell(row=current_row, column=1, value="TOTALES PARCIALES").font = Font(bold=True)
            # Get column indices for totals
            col_idx_parking = data_table_headers.index("PARKING") + 1
            col_idx_taxi = data_table_headers.index("TAXI") + 1
            col_idx_km = data_table_headers.index("KM") + 1 # This is the sum of KMs, not importe KMs
            col_idx_importe_kms = data_table_headers.index("IMPORTE KMs") + 1
            col_idx_avion = data_table_headers.index("AVION/HOTEL/COCHE") + 1
            col_idx_hotel = data_table_headers.index("HOTEL") + 1
            col_idx_almuerzo = data_table_headers.index("ALMUERZO") + 1
            col_idx_cena = data_table_headers.index("CENA") + 1
            col_idx_varios = data_table_headers.index("VARIOS") + 1
            col_idx_total_diario_sum = data_table_headers.index("TOTAL DIARIO") + 1

            partial_totals_map = [
                (col_idx_parking, totals_accumulators["PARKING"]),
                (col_idx_taxi, totals_accumulators["TAXI"]),
                (col_idx_km, totals_accumulators["KM"]),
                (col_idx_importe_kms, totals_accumulators["IMPORTE KMs"]),
                (col_idx_avion, totals_accumulators["AVION/HOTEL/COCHE"]),
                (col_idx_hotel, totals_accumulators["HOTEL"]),
                (col_idx_almuerzo, totals_accumulators["ALMUERZO"]),
                (col_idx_cena, totals_accumulators["CENA"]),
                (col_idx_varios, totals_accumulators["VARIOS"]),
                (col_idx_total_diario_sum, totals_accumulators["TOTAL DIARIO"])
            ]

            thin_border_side = Side(style='thin', color="000000")
            for col, total_val in partial_totals_map:
                cell = ws.cell(row=current_row, column=col, value=total_val)
                cell.font = Font(bold=True)
                if data_table_headers[col-1] == "KM": # KM is numeric not currency
                    cell.number_format = '0.00'
                else:
                    format_currency_cell(cell, sheet.currency)
                cell.alignment = Alignment(horizontal="right")
                cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            ws.cell(row=current_row, column=1).border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side) # Border for label

            current_row += 1 # Blank row
            current_row += 1

            # Row: SUBTOTAL
            subtotal_label_cell = ws.cell(row=current_row, column=data_table_headers.index("VARIOS") + 1, value="SUBTOTAL") # Column M
            subtotal_label_cell.font = Font(bold=True)
            subtotal_label_cell.alignment = Alignment(horizontal="right")
            subtotal_value_cell = ws.cell(row=current_row, column=data_table_headers.index("TOTAL DIARIO") + 1, value=totals_accumulators["TOTAL DIARIO"]) # Column N
            subtotal_value_cell.font = Font(bold=True)
            format_currency_cell(subtotal_value_cell, sheet.currency)
            subtotal_value_cell.alignment = Alignment(horizontal="right")
            subtotal_label_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            subtotal_value_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            current_row += 1

            # Row: ANTICIPO
            anticipo_val = sheet.anticipo if sheet.anticipo is not None else 0.0
            anticipo_label_cell = ws.cell(row=current_row, column=data_table_headers.index("VARIOS") + 1, value="ANTICIPO") # Column M
            anticipo_label_cell.font = Font(bold=True)
            anticipo_label_cell.alignment = Alignment(horizontal="right")
            anticipo_value_cell = ws.cell(row=current_row, column=data_table_headers.index("TOTAL DIARIO") + 1, value=anticipo_val) # Column N
            anticipo_value_cell.font = Font(bold=True)
            format_currency_cell(anticipo_value_cell, sheet.currency)
            anticipo_value_cell.alignment = Alignment(horizontal="right")
            anticipo_label_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            anticipo_value_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            current_row += 1

            # Row: DEVOLUCION - Assuming 'devolucion' field will be added to ActualExpenseSheet model
            devolucion_val = getattr(sheet, 'devolucion', 0.0) if getattr(sheet, 'devolucion', None) is not None else 0.0
            devolucion_label_cell = ws.cell(row=current_row, column=data_table_headers.index("VARIOS") + 1, value="DEVOLUCION") # Column M
            devolucion_label_cell.font = Font(bold=True)
            devolucion_label_cell.alignment = Alignment(horizontal="right")
            devolucion_value_cell = ws.cell(row=current_row, column=data_table_headers.index("TOTAL DIARIO") + 1, value=devolucion_val) # Column N
            devolucion_value_cell.font = Font(bold=True)
            format_currency_cell(devolucion_value_cell, sheet.currency)
            devolucion_value_cell.alignment = Alignment(horizontal="right")
            devolucion_label_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            devolucion_value_cell.border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            current_row += 1

            # Row: TOTAL
            final_total_val = totals_accumulators["TOTAL DIARIO"] + anticipo_val - devolucion_val
            total_label_cell = ws.cell(row=current_row, column=data_table_headers.index("VARIOS") + 1, value="TOTAL") # Column M
            total_label_cell.font = Font(bold=True)
            total_label_cell.alignment = Alignment(horizontal="right")
            total_value_cell = ws.cell(row=current_row, column=data_table_headers.index("TOTAL DIARIO") + 1, value=final_total_val) # Column N
            total_value_cell.font = Font(bold=True)
            format_currency_cell(total_value_cell, sheet.currency)
            total_value_cell.alignment = Alignment(horizontal="right")
            double_top_border = Border(left=thin_border_side, right=thin_border_side, top=Side(style='double', color="000000"), bottom=thin_border_side)
            total_label_cell.border = double_top_border
            total_value_cell.border = double_top_border

            # Set print area and other print settings if needed
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True


        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Sanitize sheet name (already done for worksheet title, reuse if suitable or re-sanitize)
        # Use a robust sanitization for filenames, similar to sanitize_storage_key but allow more flexibility if needed
        # For simplicity, reusing sanitize_storage_key for the sheet name part of the filename.
        sanitized_sheet_name_for_file = sanitize_storage_key(sheet.name if sheet.name else "HojaDeGastos")
        if not sanitized_sheet_name_for_file: # Ensure not empty
            sanitized_sheet_name_for_file = "HojaDeGastos"

        current_date_str = datetime.datetime.now().strftime("%Y_%m_%d")
        file_name = f"{sanitized_sheet_name_for_file}_{current_date_str}.xlsx"
        
        encoded_content = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
        
        print(f"[BASE64_EXPORT_DEBUG] Successfully generated Excel with real data and encoded content for: {sheet_id}. Filename: {file_name}")

        return ExportResponse(
            file_name=file_name,
            file_content_base64=encoded_content
        )

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        print(f"[BASE64_EXPORT_DEBUG] Unexpected error in export_expense_sheet_to_excel: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e)) from e

# --- New Endpoint to Export Receipts as ZIP ---
@router.get("/expense-sheet/{sheet_id}/receipts-zip", tags=["stream"])
async def export_expense_sheet_receipts_zip(sheet_id: str, user: AuthorizedUser):
    print(f"[RECEIPT_ZIP_EXPORT] User {user.sub} requested ZIP for sheet {sheet_id}")

    # 1. Get Expense Sheet Data
    try:
        storage_key = get_expense_sheet_storage_key(sheet_id)
        sheet_data_from_storage = db.storage.json.get(storage_key, default=None)
        if sheet_data_from_storage is None:
            print(f"[RECEIPT_ZIP_EXPORT] Expense sheet {sheet_id} not found at {storage_key}")
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Expense sheet {sheet_id} not found")

        loaded_dict = json.loads(sheet_data_from_storage) if isinstance(sheet_data_from_storage, str) else sheet_data_from_storage
        if not isinstance(loaded_dict, dict):
             print(f"[RECEIPT_ZIP_EXPORT] Failed to load sheet data {sheet_id} into a dictionary.")
             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to load sheet data into a dictionary.")
        
        sheet = ActualExpenseSheet(**loaded_dict)
        print(f"[RECEIPT_ZIP_EXPORT] Loaded sheet '{sheet.name}' with {len(sheet.entries)} entries.")

    except HTTPException:
        raise
    except Exception as e:
        print(f"[RECEIPT_ZIP_EXPORT] Error loading expense sheet {sheet_id}: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error loading expense sheet: {e}")

    # 2. Authenticate with Google Drive
    token_storage_key = f"google_tokens_{user.sub}"
    token_info = db.storage.json.get(token_storage_key, default=None)
    print(f"[RECEIPT_ZIP_EXPORT_DEBUG] User {user.sub} - Loaded token_info for key '{token_storage_key}': {token_info}")

    if not token_info or not token_info.get("refresh_token"):
        print(f"[RECEIPT_ZIP_EXPORT] Refresh token not found for user {user.sub} (key: {token_storage_key})")
        if token_info: print(f"[RECEIPT_ZIP_EXPORT] Content of {token_storage_key}: {token_info}")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Google Drive refresh token not found. Please re-authenticate.")

    google_client_id = os.environ.get("GOOGLE_CLIENT_ID")
    google_client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")

    if mode == Mode.DEV:
        try:
            # 'import databutton as db' was removed from here to avoid UnboundLocalError
            # The module-level import will be used.
            if not google_client_id:
                # Ensure db is available from module-level import for secrets fallback
                if 'db' in globals() or 'db' in locals(): # Defensive check
                    google_client_id = db.secrets.get("GOOGLE_CLIENT_ID")
                    if google_client_id: print("INFO: GOOGLE_CLIENT_ID (export_service) loaded from db.secrets (fallback).")
                else:
                    print("ERROR: 'db' module not available for GOOGLE_CLIENT_ID fallback in export_service")
            if not google_client_secret:
                if 'db' in globals() or 'db' in locals(): # Defensive check
                    google_client_secret = db.secrets.get("GOOGLE_CLIENT_SECRET")
                    if google_client_secret: print("INFO: GOOGLE_CLIENT_SECRET (export_service) loaded from db.secrets (fallback).")
                else:
                    print("ERROR: 'db' module not available for GOOGLE_CLIENT_SECRET fallback in export_service")
        except ImportError:
            print("ERROR: ImportError during fallback secret loading in export_service, db SDK might not be available as expected.")
            pass # Databutton SDK not available or other import error


    if not google_client_id or not google_client_secret:
        print("[RECEIPT_ZIP_EXPORT] Google client ID or secret not configured.")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Server configuration error for Google authentication.")

    try:
        credentials = Credentials(
            token=None, 
            refresh_token=token_info["refresh_token"],
            token_uri='https://oauth2.googleapis.com/token',
            client_id=google_client_id,
            client_secret=google_client_secret
            # Do not specify scopes here when using a refresh_token that already has them.
            # The refresh_token itself is bound to the scopes granted when it was issued.
            # Explicitly requesting a different subset like ["https://www.googleapis.com/auth/drive.readonly"]
            # can cause 'invalid_scope' if it doesn't perfectly match or isn't a pre-approved subset rule.
        )
        if not token_info.get("scopes") or "https://www.googleapis.com/auth/drive.file" not in token_info.get("scopes", []):
            # If the stored token_info doesn't explicitly list drive.file scope (it should from the logs),
            # or as a fallback, we might need to add it if Google's library requires it.
            # However, for refresh, it's best to rely on the scopes embedded in the refresh_token itself.
            # Let's try without first. If issues persist, this is a place to reconsider.
            # credentials.scopes = ["https://www.googleapis.com/auth/drive.file"] # Or drive.readonly if that's the minimum needed
            pass
        credentials.refresh(GoogleAuthRequest()) # Use aliased Request
        drive_service = build('drive', 'v3', credentials=credentials)
        print(f"[RECEIPT_ZIP_EXPORT] Google Drive service built successfully for user {user.sub}")
    except Exception as cred_err:
        print(f"[RECEIPT_ZIP_EXPORT] Failed to refresh Google credentials for user {user.sub}: {cred_err}")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail=f"Failed to obtain Google Drive access: {str(cred_err)}")

    # 3. Download Tickets and Create ZIP
    zip_buffer = io.BytesIO()
    downloaded_files_count = 0
    sane_sheet_name_for_zip_folder = sanitize_storage_key(sheet.name) if sheet.name else f"Hoja_{sheet_id[:8]}"
    if not sane_sheet_name_for_zip_folder: # Ensure it's not empty
        sane_sheet_name_for_zip_folder = f"Hoja_{sheet_id[:8]}"

    try:
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for entry_idx, entry in enumerate(sheet.entries):
                if entry.receipt_google_drive_id and entry.receipt_google_drive_file_name:
                    print(f"[RECEIPT_ZIP_EXPORT] Processing entry {entry.id} (idx {entry_idx}), Drive ID: {entry.receipt_google_drive_id}, Filename: {entry.receipt_google_drive_file_name}")
                    try:
                        request_dl = drive_service.files().get_media(fileId=entry.receipt_google_drive_id)
                        file_content_buffer = io.BytesIO()
                        downloader = MediaIoBaseDownload(file_content_buffer, request_dl)
                        
                        done = False
                        while not done:
                            status_download, done = downloader.next_chunk()
                            if status_download:
                                print(f"[RECEIPT_ZIP_EXPORT] Download progress for {entry.receipt_google_drive_file_name}: {int(status_download.progress() * 100)}%")
                        
                        sane_entry_filename = sanitize_storage_key(entry.receipt_google_drive_file_name)
                        if not sane_entry_filename:
                             _, sane_entry_filename_ext = os.path.splitext(entry.receipt_google_drive_file_name)
                             sane_entry_filename = f"ticket_{entry.id}{sane_entry_filename_ext if sane_entry_filename_ext else '.dat'}"

                        filename_in_zip = f"{sane_sheet_name_for_zip_folder}/{sane_entry_filename}"
                        
                        zip_file.writestr(filename_in_zip, file_content_buffer.getvalue())
                        downloaded_files_count += 1
                        print(f"[RECEIPT_ZIP_EXPORT] Added '{filename_in_zip}' to ZIP.")
                    
                    except HttpError as drive_err:
                        print(f"[RECEIPT_ZIP_EXPORT] Google Drive API error downloading file ID {entry.receipt_google_drive_id} ('{entry.receipt_google_drive_file_name}'): {drive_err}. Content: {drive_err.content}")
                        error_filename = f"{sane_sheet_name_for_zip_folder}/ERROR_DOWNLOADING_{sanitize_storage_key(entry.receipt_google_drive_file_name if entry.receipt_google_drive_file_name else f'ticket_id_{entry.receipt_google_drive_id}')}.txt"
                        zip_file.writestr(error_filename, f"Could not download file: {entry.receipt_google_drive_file_name} (Drive ID: {entry.receipt_google_drive_id}). Error: {drive_err.resp.status} - {drive_err.content.decode('utf-8', 'ignore')}")
                    except Exception as e_download:
                        print(f"[RECEIPT_ZIP_EXPORT] General error downloading/zipping file ID {entry.receipt_google_drive_id} ('{entry.receipt_google_drive_file_name}'): {e_download}")
                        error_filename_general = f"{sane_sheet_name_for_zip_folder}/ERROR_PROCESSING_{sanitize_storage_key(entry.receipt_google_drive_file_name if entry.receipt_google_drive_file_name else f'ticket_id_{entry.receipt_google_drive_id}')}.txt"
                        zip_file.writestr(error_filename_general, f"Could not process file: {entry.receipt_google_drive_file_name} (Drive ID: {entry.receipt_google_drive_id}). Error: {str(e_download)}")

                elif entry.receipt_google_drive_id and not entry.receipt_google_drive_file_name:
                    print(f"[RECEIPT_ZIP_EXPORT] Skipping entry {entry.id} (Drive ID: {entry.receipt_google_drive_id}) due to missing receipt_google_drive_file_name.")
                    missing_name_filename = f"{sane_sheet_name_for_zip_folder}/SKIPPED_MISSING_FILENAME_ID_{entry.receipt_google_drive_id}.txt"
                    zip_file.writestr(missing_name_filename, f"Skipped downloading file with Drive ID: {entry.receipt_google_drive_id} because its name was not recorded in the expense entry.")

            if downloaded_files_count == 0:
                 print(f"[RECEIPT_ZIP_EXPORT] No receipts were successfully downloaded for sheet {sheet_id}. Zip might be empty or contain only error/skipped messages.")
                 if not sheet.entries or not any(e.receipt_google_drive_id for e in sheet.entries):
                     zip_file.writestr(f"{sane_sheet_name_for_zip_folder}/NO_RECEIPTS_FOUND.txt", "No receipts with Google Drive links were found in this expense sheet.")

    except zipfile.BadZipFile as bzf_err:
        print(f"[RECEIPT_ZIP_EXPORT] Error creating ZIP file for sheet {sheet_id}: {bzf_err}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error creating ZIP archive: {bzf_err}")
    except Exception as e_zip: 
        print(f"[RECEIPT_ZIP_EXPORT] General error during ZIP creation process for sheet {sheet_id}: {e_zip}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error processing ZIP export: {e_zip}")

    # 4. Prepare and Return StreamingResponse
    zip_buffer.seek(0)
    zip_file_name_final = f"{sane_sheet_name_for_zip_folder}_Tickets.zip"
    
    print(f"[RECEIPT_ZIP_EXPORT] Preparing StreamingResponse with filename: '{zip_file_name_final}', {downloaded_files_count} files in ZIP.")

    return StreamingResponse(
        iter([zip_buffer.getvalue()]), 
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=\"{zip_file_name_final}\""} # Ensure filename in quotes
    )

# Health check for the service router

@router.get("/health_check", tags=["Service Health"])
async def service_health_check():
    return {"status": "Export service router is active and healthy."}

